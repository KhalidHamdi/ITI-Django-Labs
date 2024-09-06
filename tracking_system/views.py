from django.shortcuts import render, redirect
from openpyxl import load_workbook

EXCEL_FILE = 'tracking_system/data/orders.xlsx'

# Function to read data from Excel
def read_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        order = {
            'orderID': row[0],
            'orderName': row[1],
            'username': row[2],
            'status': row[3],
        }
        orders.append(order)
    return orders

# View to handle tracking system and adding products to cart
def tracking(request):
    orders = read_excel()

    if request.method == 'POST':
        # Get selected orders from the form
        selected_order_ids = request.POST.getlist('selected_orders')

        # Filter selected orders to add them to the cart (or other logic)
        selected_orders = [order for order in orders if str(order['orderID']) in selected_order_ids]

        # Here you would save selected orders to the user's session or cart
        request.session['cart'] = selected_orders

        # Redirect user to cart page or tracking confirmation
        return redirect('cart')  # Make sure 'cart' view exists

    return render(request, 'tracking.html', {'orders': orders})


